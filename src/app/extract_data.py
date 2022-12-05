import os
from openpyxl import load_workbook
from config import constants as const


class ExcelData:
    def __init__(self, filepath):
        self.filepath = filepath

    def extract_data(self):
        wb = load_workbook(self.filepath)
        ws = wb[wb.sheetnames[const.EXCEL_SHEET_NUMBER]]
        data = list(ws.rows)[const.EXCEL_FIRST_ROW:]

        return data
